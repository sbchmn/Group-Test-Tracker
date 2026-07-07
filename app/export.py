"""
Export / Backup feature for Group Test Manager.
Generates a .xlsx file that closely mirrors the structure and columns of the original
"MassPurity & Endo" group test spreadsheet, with live data + Excel formulas for calculations.
Uses openpyxl per project standards (professional formatting, formulas where logical).
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import GroupTest, Participation


def generate_test_export(test: GroupTest) -> BytesIO:
    """
    Create a professional .xlsx backup/export of a single GroupTest.
    Includes metadata, lab costs, full participants table, INPUTS, and CALCULATIONS with formulas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MassPurity & Endo"

    # Styles
    header_font = Font(bold=True, size=12, name='Arial')
    title_font = Font(bold=True, size=14, name='Arial')
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '"$"#,##0.00'
    center_align = Alignment(horizontal='center', vertical='center')

    # === HEADER / METADATA (matching original layout) ===
    ws.merge_cells('A1:N1')
    ws['A1'] = f"GROUP TEST EXPORT: {test.title}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = "GROUP START DATE"
    ws['B3'] = test.start_date
    ws['B3'].number_format = 'YYYY-MM-DD'

    ws['A4'] = "Vendor"
    ws['B4'] = test.vendor or ""
    ws['D4'] = "LAB / PROVIDER"
    ws['E4'] = test.lab_name or ""

    ws['A5'] = "Batch Number"
    ws['B5'] = test.batch_number or ""
    ws['D5'] = "LAB TEST NAME"
    ws['E5'] = "PRICE"
    ws['F5'] = "# VIALS NEEDED"

    ws['A6'] = "Compound"
    ws['B6'] = test.compound or ""
    ws['D6'] = "STATUS"
    ws['E6'] = test.status.upper()

    ws['A7'] = "Size"
    ws['B7'] = test.size or ""
    ws['D7'] = "TOTAL LAB COST"
    ws['E7'] = test.total_lab_cost or 0
    ws['E7'].number_format = currency_format

    ws['A8'] = "SHIPMENT TO LAB / COST"
    ws['B8'] = test.shipping_cost or 0
    ws['B8'].number_format = currency_format

    ws['A9'] = "ORDER NUMBER"
    ws['B9'] = test.order_number or ""
    ws['D9'] = "QUOTE NUMBER"
    ws['E9'] = test.quote_number or ""

    ws['A10'] = "RESULTS LINK"
    ws['B10'] = test.results_link or ""

    if test.lab_test_details:
        for idx, item in enumerate(test.lab_test_details, start=1):
            row = 10 + idx
            ws.cell(row=row, column=4, value=item.get('name') or '')
            ws.cell(row=row, column=5, value=item.get('price') or 0)
            ws.cell(row=row, column=5).number_format = currency_format
            ws.cell(row=row, column=6, value=item.get('vials_needed') or 0)

    if test.status == 'closed' and test.results_link:
        ws['A11'] = "RESULTS LINK (Closed)"
        ws.merge_cells('B11:E11')
        ws['B11'] = test.results_link
        ws['B11'].font = Font(color="0563C1", underline="single")

    # === PARTICIPANTS TABLE ===
    start_row = 15
    ws.merge_cells(f'A{start_row}:O{start_row}')
    ws[f'A{start_row}'] = "PARTICIPANTS"
    ws[f'A{start_row}'].font = header_font
    ws[f'A{start_row}'].fill = section_fill

    headers = [
        "#", "NAME", "TG Username", "Verified", "ACTIVE", "Order Status",
        "US Based", "Vial Donor", "State", "Pay Vial Collector", "Pay Lab",
        "Paid Lab", "Amount Owed", "Amount Paid", "Notes / Approved"
    ]
    header_row = start_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.border = thin_border
        cell.alignment = center_align

    # Data rows
    participations = test.participations.order_by(Participation.requested_at).all()
    data_start = header_row + 1
    for idx, p in enumerate(participations, 1):
        row = data_start + idx - 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=p.name or p.user.username).border = thin_border
        ws.cell(row=row, column=3, value=p.tg_username or p.user.tg_username or "").border = thin_border
        ws.cell(row=row, column=4, value="Yes" if p.verified else "No").border = thin_border
        ws.cell(row=row, column=5, value="Yes" if p.active else "No").border = thin_border
        ws.cell(row=row, column=6, value=p.order_status).border = thin_border
        ws.cell(row=row, column=7, value="Yes" if p.us_based else "No").border = thin_border
        ws.cell(row=row, column=8, value="Yes" if p.vial_donor else "No").border = thin_border
        ws.cell(row=row, column=9, value=p.state or "").border = thin_border
        ws.cell(row=row, column=10, value="Yes" if p.pay_vial_collector else "No").border = thin_border
        ws.cell(row=row, column=11, value="Yes" if p.pay_lab else "No").border = thin_border
        ws.cell(row=row, column=12, value="Yes" if p.paid_lab else "No").border = thin_border
        ws.cell(row=row, column=13, value=p.amount_owed).number_format = currency_format
        ws.cell(row=row, column=13).border = thin_border
        ws.cell(row=row, column=14, value=p.amount_paid).number_format = currency_format
        ws.cell(row=row, column=14).border = thin_border
        approved_str = "Approved" if p.approved else "Pending"
        ws.cell(row=row, column=15, value=f"{approved_str} | {p.notes or ''}").border = thin_border

    data_end = data_start + len(participations) - 1 if participations else data_start

    # === INPUTS SECTION ===
    input_row = data_end + 3
    ws.merge_cells(f'A{input_row}:E{input_row}')
    ws[f'A{input_row}'] = "INPUTS"
    ws[f'A{input_row}'].font = header_font
    ws[f'A{input_row}'].fill = input_fill

    inputs = [
        ("TOTAL GROUP PARTICIPANTS", f"=COUNTA(B{data_start}:B{data_end})" if participations else 0),
        ("TOTAL DONORS", f'=COUNTIF(H{data_start}:H{data_end},"Yes")' if participations else 0),
        ("TOTAL LAB TESTS COST", test.total_lab_cost or 0),
        ("FINAL POSTAGE COST", test.shipping_cost or 0),
        ("TOTAL REFUND AMOUNT TO DONOR (per donor)", test.refund_per_donor or 0),
        ("TOTAL NON-VIAL COLLECTORS", f'=COUNTIF(H{data_start}:H{data_end},"No")' if participations else 0),
    ]

    for i, (label, value) in enumerate(inputs):
        r = input_row + 1 + i
        ws.cell(row=r, column=1, value=label).fill = input_fill
        cell = ws.cell(row=r, column=2, value=value)
        cell.fill = input_fill
        if isinstance(value, (int, float)):
            cell.number_format = currency_format

    # === CALCULATIONS SECTION (with formulas) ===
    calc_row = input_row + len(inputs) + 3
    ws.merge_cells(f'A{calc_row}:E{calc_row}')
    ws[f'A{calc_row}'] = "CALCULATIONS (Excel formulas - will update if you edit data)"
    ws[f'A{calc_row}'].font = header_font
    ws[f'A{calc_row}'].fill = calc_fill

    calcs = [
        ("TOTAL NON-DONORS", f"=B{input_row+6}"),  # references TOTAL NON-VIAL COLLECTORS
        ("BASE COST PER DONOR (approx)", f"=IF(B{input_row+2}>0,(B{input_row+3}+B{input_row+4}-B{input_row+5}*B{input_row+2})/B{input_row+2},0)"),
        ("BASE COST PER NON-DONOR (approx)", f"=IF(B{input_row+6}>0,(B{input_row+3}+B{input_row+4}+B{input_row+5}*B{input_row+2})/B{input_row+6},0)"),
        ("VIAL SHIPPER REFUND PER PERSON (example)", f"=B{input_row+5}"),
        ("EACH PERSON PAY DONOR SHIPPING (example)", f"=IF(B{input_row+1}>0,B{input_row+4}/B{input_row+1},0)"),
    ]

    for i, (label, formula) in enumerate(calcs):
        r = calc_row + 1 + i
        ws.cell(row=r, column=1, value=label).fill = calc_fill
        cell = ws.cell(row=r, column=2, value=formula)
        cell.fill = calc_fill
        cell.number_format = currency_format

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Freeze panes for participants table
    ws.freeze_panes = f'A{header_row + 1}'

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output